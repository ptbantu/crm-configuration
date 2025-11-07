#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
读取 Vendors.xlsx 文件并生成供应商组织的 SQL 插入语句
"""

import sys
import uuid
import json
from datetime import datetime

try:
    import openpyxl
except ImportError:
    print("错误: 需要安装 openpyxl 库")
    print("运行: pip install openpyxl")
    sys.exit(1)

def generate_uuid():
    """生成固定格式的 UUID（用于种子数据）"""
    return str(uuid.uuid4())

def escape_sql_string(value):
    """转义 SQL 字符串"""
    if value is None:
        return 'NULL'
    if isinstance(value, (int, float)):
        return str(value)
    if isinstance(value, bool):
        return 'TRUE' if value else 'FALSE'
    # 转义单引号
    return "'" + str(value).replace("'", "''") + "'"

def read_vendors_excel(file_path):
    """读取 Vendors.xlsx 文件"""
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active
    
    print(f"工作表: {ws.title}")
    print(f"行数: {ws.max_row}, 列数: {ws.max_column}")
    print()
    
    # 读取表头
    headers = []
    for cell in ws[1]:
        headers.append(cell.value if cell.value else f"Column{cell.column_letter}")
    
    print("表头:")
    for i, header in enumerate(headers, 1):
        print(f"  {i}: {header}")
    print()
    
    # 读取数据行
    vendors = []
    for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
        if not any(row):  # 跳过空行
            continue
        
        vendor = {}
        for i, value in enumerate(row):
            if i < len(headers):
                vendor[headers[i]] = value
        
        vendors.append(vendor)
    
    print(f"找到 {len(vendors)} 个供应商")
    print()
    
    # 显示前3条数据示例
    print("前3条数据示例:")
    for i, vendor in enumerate(vendors[:3], 1):
        print(f"\n供应商 {i}:")
        for key, value in vendor.items():
            print(f"  {key}: {value}")
    
    return headers, vendors

def generate_sql(vendors, headers):
    """生成 SQL 插入语句"""
    sql_lines = []
    
    sql_lines.append("-- ============================================================")
    sql_lines.append("-- 供应商组织种子数据")
    sql_lines.append("-- 从 Vendors.xlsx 文件生成")
    sql_lines.append("-- ============================================================")
    sql_lines.append("")
    sql_lines.append("-- 注意：使用 INSERT IGNORE 避免重复插入（基于 code 字段）")
    sql_lines.append("")
    
    for idx, vendor in enumerate(vendors, 1):
        # 生成组织编码（如果没有，使用索引）
        org_code = None
        if 'Code' in vendor and vendor['Code']:
            org_code = str(vendor['Code']).strip().upper()
        elif 'code' in vendor and vendor['code']:
            org_code = str(vendor['code']).strip().upper()
        elif 'Name' in vendor and vendor['Name']:
            # 从名称生成编码
            name = str(vendor['Name']).strip()
            org_code = name.upper().replace(' ', '_').replace('-', '_')[:20]
        else:
            org_code = f"VENDOR_{idx:03d}"
        
        # 获取组织名称
        org_name = None
        if 'Name' in vendor and vendor['Name']:
            org_name = str(vendor['Name']).strip()
        elif 'name' in vendor and vendor['name']:
            org_name = str(vendor['name']).strip()
        else:
            org_name = f"供应商 {idx}"
        
        if not org_name:
            continue
        
        # 开始生成 SQL
        sql_lines.append(f"-- 供应商 {idx}: {org_name}")
        sql_lines.append("INSERT IGNORE INTO organizations (")
        sql_lines.append("    id,")
        sql_lines.append("    name,")
        sql_lines.append("    code,")
        sql_lines.append("    organization_type,")
        sql_lines.append("    parent_id,")
        sql_lines.append("    email,")
        sql_lines.append("    phone,")
        sql_lines.append("    website,")
        sql_lines.append("    description,")
        sql_lines.append("    street,")
        sql_lines.append("    city,")
        sql_lines.append("    state_province,")
        sql_lines.append("    postal_code,")
        sql_lines.append("    country,")
        sql_lines.append("    country_code,")
        sql_lines.append("    company_size,")
        sql_lines.append("    company_nature,")
        sql_lines.append("    company_type,")
        sql_lines.append("    industry,")
        sql_lines.append("    is_active,")
        sql_lines.append("    is_locked,")
        sql_lines.append("    created_at,")
        sql_lines.append("    updated_at")
        sql_lines.append(") VALUES (")
        
        # 使用 MySQL 的 UUID() 函数自动生成 UUID
        org_id = "UUID()"
        
        # 提取字段值
        email = vendor.get('Email') or vendor.get('email') or vendor.get('Email Address')
        phone = vendor.get('Phone') or vendor.get('phone') or vendor.get('Phone Number')
        website = vendor.get('Website') or vendor.get('website') or vendor.get('Web')
        description = vendor.get('Description') or vendor.get('description') or vendor.get('Notes')
        street = vendor.get('Street') or vendor.get('street') or vendor.get('Address')
        city = vendor.get('City') or vendor.get('city')
        state = vendor.get('State') or vendor.get('state') or vendor.get('Province')
        postal = vendor.get('Postal') or vendor.get('postal') or vendor.get('Postal Code')
        country = vendor.get('Country') or vendor.get('country') or '中国'
        country_code = vendor.get('Country Code') or vendor.get('country_code') or 'CN'
        
        sql_lines.append(f"    {org_id},")  # UUID() 函数，不需要转义
        sql_lines.append(f"    {escape_sql_string(org_name)},")
        sql_lines.append(f"    {escape_sql_string(org_code)},")
        sql_lines.append("    'vendor',")
        sql_lines.append("    NULL,")
        sql_lines.append(f"    {escape_sql_string(email)},")
        sql_lines.append(f"    {escape_sql_string(phone)},")
        sql_lines.append(f"    {escape_sql_string(website)},")
        sql_lines.append(f"    {escape_sql_string(description)},")
        sql_lines.append(f"    {escape_sql_string(street)},")
        sql_lines.append(f"    {escape_sql_string(city)},")
        sql_lines.append(f"    {escape_sql_string(state)},")
        sql_lines.append(f"    {escape_sql_string(postal)},")
        sql_lines.append(f"    {escape_sql_string(country)},")
        sql_lines.append(f"    {escape_sql_string(country_code)},")
        sql_lines.append("    'medium',")  # 默认公司规模
        sql_lines.append("    'private',")  # 默认公司性质
        sql_lines.append("    'limited',")  # 默认公司类型
        sql_lines.append("    NULL,")  # 行业（待补充）
        sql_lines.append("    TRUE,")
        sql_lines.append("    FALSE,")
        sql_lines.append("    CURRENT_TIMESTAMP,")
        sql_lines.append("    CURRENT_TIMESTAMP")
        sql_lines.append(");")
        sql_lines.append("")
    
    return "\n".join(sql_lines)

if __name__ == "__main__":
    excel_file = "Vendors.xlsx"
    
    if len(sys.argv) > 1:
        excel_file = sys.argv[1]
    
    print("=" * 60)
    print("读取供应商 Excel 文件")
    print("=" * 60)
    print()
    
    try:
        headers, vendors = read_vendors_excel(excel_file)
        
        if not vendors:
            print("警告: 未找到供应商数据")
            sys.exit(1)
        
        print()
        print("=" * 60)
        print("生成 SQL 文件")
        print("=" * 60)
        print()
        
        sql_content = generate_sql(vendors, headers)
        
        output_file = "vendors_seed_data.sql"
        with open(output_file, 'w', encoding='utf-8') as f:
            f.write(sql_content)
        
        print(f"✅ SQL 文件已生成: {output_file}")
        print(f"   包含 {len(vendors)} 个供应商组织")
        print()
        print("使用方法:")
        print(f"  mysql -u user -p database < {output_file}")
        print()
        
    except Exception as e:
        print(f"❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)

